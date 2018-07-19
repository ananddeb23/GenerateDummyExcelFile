from openpyxl import Workbook
import random
import string
wb = Workbook()

ws = wb.active
rows = 20
cols = 40
lower_range_limit_of_random_number = 100.0
upper_range_limit_of_random_number = 100000.5
random_number_floating_precision = 4 
length_of_random_text = 10

def generateRandomValue():
  numtextchoice = random.randint(0,1)
  result = dict();
  # 50% chance of generating a random number
  if(numtextchoice == 0):
    result['format'] = '0' 
    result['value'] = round(random.uniform(lower_range_limit_of_random_number,upper_range_limit_of_random_number), random_number_floating_precision)
    return result
  # 50% chance of generating a random text(mix of lowercase, uppercase and digits)
  else:
    result['format'] = '@'
    result['value'] = ''.join(random.choice(string.ascii_uppercase + string.ascii_lowercase + string.digits) for _ in range(10))  
    return  result
def generateFileName():
  return 'DummyExcelWith'+str(rows)+'rows'+str(cols)+'cols'+'.xlsx'

ws = wb.active
for i in range(1, rows+1):
  for j in range(1, cols+1):
    val = generateRandomValue()
    print(val)
    ws.cell(row=i, column=j, value=val['value'])
    ws.cell(row=i,column=j).number_format = val['format']

file_name = generateFileName()
wb.save(file_name)